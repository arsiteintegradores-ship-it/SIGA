from datetime import date, datetime, timedelta
from decimal import Decimal

from django import forms
from django.contrib import admin
from django.contrib import messages
from django.contrib.admin.widgets import AdminDateWidget
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm

from .models import (
    GanadoAnimal,
    GanadoFinca,
    GanadoLote,
    GanadoRaza,
    GanadoColor,
    GanadoProductor,
    GanadoUpp,
    GanadoRegistro,
)
from .utils import normalize_identifier


admin.site.site_header = "SIGA - Administración"
admin.site.site_title = "SIGA Admin"
admin.site.index_title = ""

# =========================
# Filtros por rango (sin paquetes)
# =========================

class FechaNacimientoFilter(admin.SimpleListFilter):
    title = "Fecha nacimiento (rango)"
    parameter_name = "fn_rango"

    def lookups(self, request, model_admin):
        return (
            ("hoy", "Hoy"),
            ("30", "Últimos 30 días"),
            ("90", "Últimos 90 días"),
            ("365", "Últimos 365 días"),
            ("ant", "Más de 365 días"),
        )

    def queryset(self, request, queryset):
        val = self.value()
        if not val:
            return queryset

        today = timezone.localdate()
        if val == "hoy":
            return queryset.filter(fecha_nacimiento=today)
        if val in ("30", "90", "365"):
            days = int(val)
            return queryset.filter(fecha_nacimiento__gte=today - timedelta(days=days))
        if val == "ant":
            return queryset.filter(fecha_nacimiento__lt=today - timedelta(days=365))
        return queryset


class PesoNacimientoFilter(admin.SimpleListFilter):
    title = "Peso nacimiento (kg)"
    parameter_name = "pn_rango"

    def lookups(self, request, model_admin):
        return (
            ("0-25", "0 a 25"),
            ("25-35", "25 a 35"),
            ("35-45", "35 a 45"),
            ("45+", "45+"),
            ("null", "Sin dato"),
        )

    def queryset(self, request, queryset):
        val = self.value()
        if not val:
            return queryset
        if val == "null":
            return queryset.filter(peso_nacimiento__isnull=True)
        if val == "45+":
            return queryset.filter(peso_nacimiento__gte=45)
        a, b = val.split("-")
        return queryset.filter(peso_nacimiento__gte=float(a), peso_nacimiento__lt=float(b))


class PesoDesteteFilter(admin.SimpleListFilter):
    title = "Peso destete (kg)"
    parameter_name = "pd_rango"

    def lookups(self, request, model_admin):
        return (
            ("0-150", "0 a 150"),
            ("150-200", "150 a 200"),
            ("200-250", "200 a 250"),
            ("250+", "250+"),
            ("null", "Sin dato"),
        )

    def queryset(self, request, queryset):
        val = self.value()
        if not val:
            return queryset
        if val == "null":
            return queryset.filter(peso_destete__isnull=True)
        if val == "250+":
            return queryset.filter(peso_destete__gte=250)
        a, b = val.split("-")
        return queryset.filter(peso_destete__gte=float(a), peso_destete__lt=float(b))


class SinigaVacioFilter(admin.SimpleListFilter):
    title = "By Siniga"
    parameter_name = "siniga_vacio"

    def lookups(self, request, model_admin):
        return (("sin", "Sin Siniga"),)

    def queryset(self, request, queryset):
        if self.value() == "sin":
            return queryset.filter(Q(id_siniga__isnull=True) | Q(id_siniga=""))
        return queryset


# =========================
# Exportar a Excel (XLSX)
# =========================
def export_animales_xlsx(modeladmin, request, queryset):
    """
    Exporta animales seleccionados a XLSX.
    Incluye Edad (texto) y Etapa de desarrollo.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse(
            "Falta openpyxl. Instala con: pip install openpyxl",
            content_type="text/plain",
            status=500,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Animales"

    headers = [
        "ID", "Id interno", "Id SINIGA", "Nombre", "Sexo",
        "Raza", "Color", "Finca", "Lote", "Estado",
        "Fecha nac",
        "Edad (Años, Meses, Días)", "Etapa de desarrollo",
        "Peso nac", "Peso destete",
        "Padre", "Madre", "Productor", "UPP", "Notas",
        "Creado", "Actualizado",
    ]
    ws.append(headers)

    queryset = queryset.select_related(
        "raza", "color", "finca", "lote", "padre", "madre", "productor", "upp"
    )

    fecha_col = headers.index("Fecha nac") + 1

    for a in queryset:
        edad_txt = str(getattr(a, "edad", "") or "")
        etapa_txt = str(getattr(a, "etapa_desarrollo", "") or "")
        fecha_value = a.fecha_nacimiento
        fecha_date = None
        if isinstance(fecha_value, datetime):
            fecha_date = fecha_value.date()
        elif isinstance(fecha_value, date):
            fecha_date = fecha_value
        elif isinstance(fecha_value, str) and fecha_value:
            for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                try:
                    fecha_date = datetime.strptime(fecha_value, fmt).date()
                    break
                except ValueError:
                    continue

        ws.append([
            a.id,
            a.id_interno,
            a.id_siniga,
            a.nombre_bov,
            a.sexo,
            str(a.raza) if a.raza_id else "",
            str(a.color) if a.color_id else "",
            str(a.finca) if a.finca_id else "",
            str(a.lote) if a.lote_id else "",
            a.estado,
            fecha_date or "",
            edad_txt,
            etapa_txt,
            float(a.peso_nacimiento) if a.peso_nacimiento is not None else "",
            float(a.peso_destete) if a.peso_destete is not None else "",
            str(a.padre) if a.padre_id else "",
            str(a.madre) if a.madre_id else "",
            str(a.productor) if a.productor_id else "",
            str(a.upp) if a.upp_id else "",
            a.notas or "",
            a.created_at.isoformat() if a.created_at else "",
            a.updated_at.isoformat() if a.updated_at else "",
        ])
        if fecha_date:
            cell = ws.cell(row=ws.max_row, column=fecha_col)
            cell.number_format = "dd/mm/yyyy"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="siga_animales.xlsx"'
    wb.save(response)
    return response

export_animales_xlsx.short_description = "Exportar animales seleccionados a Excel (XLSX)"


def imprimir_reporte_animales_pdf(modeladmin, request, queryset):
    """
    Genera un PDF listo para imprimir de los animales seleccionados.
    Incluye Edad (texto) y Etapa.
    """
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_animales.pdf"'

    c = canvas.Canvas(response, pagesize=letter)

    headers = [
        "Interno",
        "Siniga",
        "Nombre",
        "Raza",
        "Sx",
        "F. Nac.",
        "Edad",
        "Etapa",
        "Productor",
        "Padre",
        "Madre",
        "Estado",
    ]
    min_widths = [55, 110, 110, 80, 35, 90, 190, 120, 170, 90, 90, 70]
    header_font = "Helvetica-Bold"
    header_size = 9
    body_font = "Helvetica"
    body_size = 8.5

    margin_x = 1.5 * cm
    margin_top = 2 * cm
    margin_bottom = 2 * cm

    def required_width():
        return sum(
            max(min_w, c.stringWidth(h, header_font, header_size) + 6)
            for h, min_w in zip(headers, min_widths)
        )

    width, height = letter
    available_width = width - (2 * margin_x)
    if required_width() > available_width:
        width, height = landscape(letter)
        c.setPageSize((width, height))
        available_width = width - (2 * margin_x)

    base_widths = [
        max(min_w, c.stringWidth(h, header_font, header_size) + 6)
        for h, min_w in zip(headers, min_widths)
    ]
    base_total = sum(base_widths) or 1
    scale = available_width / base_total
    cols = [w * scale for w in base_widths]

    queryset = queryset.select_related("finca", "raza", "padre", "madre", "productor")
    first_animal = queryset.first()
    finca_name = str(first_animal.finca) if first_animal and first_animal.finca_id else ""
    date_str = timezone.localdate().strftime("%d/%m/%Y")

    def fit_text(text, max_width):
        text = str(text or "")
        if c.stringWidth(text, body_font, body_size) <= max_width:
            return text
        ellipsis = "..."
        max_width = max(0, max_width - c.stringWidth(ellipsis, body_font, body_size))
        for i in range(len(text), 0, -1):
            if c.stringWidth(text[:i], body_font, body_size) <= max_width:
                return text[:i] + ellipsis
        return ellipsis

    def draw_header():
        c.setFont("Helvetica-Bold", 11)
        c.setFillColor(colors.black)
        c.drawString(margin_x, height - margin_top, finca_name)
        c.setFont("Helvetica", 9)
        c.drawRightString(width - margin_x, height - margin_top, date_str)
        c.setStrokeColor(colors.HexColor("#9A9A9A"))
        c.line(margin_x, height - margin_top - 4, width - margin_x, height - margin_top - 4)

    def draw_footer():
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor("#666666"))
        c.drawRightString(width - margin_x, 1.5 * cm, f"Pagina {c.getPageNumber()}")

    def draw_table_header(y_pos):
        header_height = 14
        c.setFillColor(colors.HexColor("#F0F0F0"))
        c.rect(margin_x, y_pos - header_height + 2, available_width, header_height, fill=1, stroke=0)
        c.setFillColor(colors.black)
        c.setFont(header_font, header_size)
        x = margin_x
        for h, w in zip(headers, cols):
            c.drawString(x + 2, y_pos, h)
            x += w
        c.setStrokeColor(colors.HexColor("#B3B3B3"))
        c.line(margin_x, y_pos - header_height + 2, width - margin_x, y_pos - header_height + 2)
        return y_pos - header_height - 4

    y = height - margin_top - 0.7 * cm
    draw_header()
    y = draw_table_header(y)

    line_h = 12
    c.setFont(body_font, body_size)
    c.setStrokeColor(colors.HexColor("#E0E0E0"))

    def nueva_pagina():
        nonlocal y
        draw_footer()
        c.showPage()
        draw_header()
        y = height - margin_top - 0.7 * cm
        y = draw_table_header(y)
        c.setFont(body_font, body_size)
        c.setStrokeColor(colors.HexColor("#E0E0E0"))

    total_animales = 0
    for a in queryset:
        if y < margin_bottom + line_h:
            nueva_pagina()

        total_animales += 1
        fila = [
            str(a.id_interno or ""),
            str(a.id_siniga) if a.id_siniga else "Sin Siniga",
            str(a.nombre_bov or ""),
            str(a.raza) if a.raza_id else "",
            str(a.sexo or ""),
            a.fecha_nacimiento.strftime("%d/%m/%Y") if a.fecha_nacimiento else "",
            str(getattr(a, "edad", "") or ""),
            str(getattr(a, "etapa_desarrollo", "") or ""),
            str(a.productor) if a.productor_id else "",
            str(a.padre) if a.padre_id else "",
            str(a.madre) if a.madre_id else "",
            str(a.estado or ""),
        ]

        x = margin_x
        for val, w in zip(fila, cols):
            txt = fit_text(val, w - 4)
            c.drawString(x + 2, y, txt)
            x += w

        c.line(margin_x, y - 3, width - margin_x, y - 3)
        y -= line_h

    if y < margin_bottom + (2 * line_h):
        nueva_pagina()

    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(colors.black)
    y -= line_h
    c.drawString(margin_x, y, f"Total animales: {total_animales}")
    y -= line_h

    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(colors.HexColor("#666666"))
    c.drawString(margin_x, 1.5 * cm, "Generado por SIGA")
    draw_footer()
    c.save()
    return response

imprimir_reporte_animales_pdf.short_description = "Imprimir reporte PDF (seleccionados)"


# =========================
# Acciones rápidas (estado)
# =========================
def set_estado(estado):
    def _action(modeladmin, request, queryset):
        queryset.update(estado=estado)

    _action.short_description = f"Marcar como {estado}"
    _action.__name__ = f"marcar_{estado.lower()}"
    return _action


class GanadoAnimalAdminForm(forms.ModelForm):
    # Forzamos el formato de captura a DD/MM/AAAA
    fecha_nacimiento = forms.DateField(
        required=False,
        input_formats=["%d/%m/%Y"],
        widget=AdminDateWidget(
            format="%d/%m/%Y",
            attrs={"placeholder": "DD/MM/AAAA", "data-date-format": "dd/mm/yyyy"},
        ),
        help_text="Formato: DD/MM/AAAA (ej: 17/01/2026)",
    )

    class Media:
        js = ("ganado/js/date_mask.js",)

    
    class Meta:
        model = GanadoAnimal
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.fields["productor"].required = True
        self.fields["raza"].required = True
        self.fields["fecha_nacimiento"].required = True
        self.fields["peso_nacimiento"].required = True

        if self.instance and getattr(self.instance, "finca_id", None):
            self.fields["lote"].queryset = GanadoLote.objects.filter(finca_id=self.instance.finca_id)
        else:
            self.fields["lote"].queryset = GanadoLote.objects.none()

    def clean_id_interno(self):
        return normalize_identifier(self.cleaned_data.get("id_interno"))

    def clean_id_siniga(self):
        return normalize_identifier(self.cleaned_data.get("id_siniga"))

    def clean(self):
        cleaned_data = super().clean()
        qs = GanadoAnimal.objects.all()
        if self.instance and self.instance.pk:
            qs = qs.exclude(pk=self.instance.pk)

        id_interno = cleaned_data.get("id_interno")
        if id_interno and qs.filter(id_interno=id_interno).exists():
            self.add_error("id_interno", "El ID interno ya existe.")

        id_siniga = cleaned_data.get("id_siniga")
        if id_siniga and qs.filter(id_siniga=id_siniga).exists():
            self.add_error("id_siniga", "El ID Siniga ya existe.")

        nombre_bov = cleaned_data.get("nombre_bov")
        if nombre_bov and qs.filter(nombre_bov__iexact=nombre_bov).exists():
            self.add_error("nombre_bov", "El nombre ya existe.")

        return cleaned_data


# =========================
# Admins de catálogos
# =========================
@admin.register(GanadoFinca)
class GanadoFincaAdmin(admin.ModelAdmin):
    list_display = ("id", "nombre", "ubicacion", "telefono", "activo")
    search_fields = ("nombre", "ubicacion", "propietario", "telefono")
    list_filter = ("activo",)
    ordering = ("nombre",)


@admin.register(GanadoLote)
class GanadoLoteAdmin(admin.ModelAdmin):
    list_display = ("id", "nombre", "finca", "activo")
    search_fields = ("nombre", "descripcion", "finca__nombre")
    list_filter = ("activo", "finca")
    ordering = ("finca__nombre", "nombre")
    autocomplete_fields = ("finca",)


@admin.register(GanadoRaza)
class GanadoRazaAdmin(admin.ModelAdmin):
    list_display = ("id", "raza")
    search_fields = ("raza",)
    ordering = ("raza",)


@admin.register(GanadoColor)
class GanadoColorAdmin(admin.ModelAdmin):
    list_display = ("id", "color")
    search_fields = ("color",)
    ordering = ("color",)


@admin.register(GanadoProductor)
class GanadoProductorAdmin(admin.ModelAdmin):
    list_display = ("id", "nombre", "apellido_paterno", "apellido_materno", "telefono", "email", "activo")
    search_fields = ("nombre", "apellido_paterno", "apellido_materno", "telefono", "email")
    list_filter = ("activo",)
    ordering = ("nombre", "apellido_paterno", "apellido_materno")


@admin.register(GanadoUpp)
class GanadoUppAdmin(admin.ModelAdmin):
    list_display = ("id", "clave", "finca", "productor")
    search_fields = ("clave", "finca__nombre", "productor__nombre", "productor__apellido_paterno", "productor__apellido_materno")
    list_filter = ("finca", "productor")
    ordering = ("finca__nombre", "clave")
    autocomplete_fields = ("finca", "productor")


@admin.register(GanadoRegistro)
class GanadoRegistroAdmin(admin.ModelAdmin):
    list_display = ("id", "id_bovino", "id_madre", "id_padre")
    search_fields = ("id_bovino", "id_madre", "id_padre", "id_abuelo", "id_abuela")
    ordering = ("-id",)


# =========================
# Admin principal: ANIMAL
# =========================
@admin.register(GanadoAnimal)
class GanadoAnimalAdmin(admin.ModelAdmin):
    form = GanadoAnimalAdminForm

    list_display = (
        "id_interno",
        "id_siniga_admin",
        "nombre_bov",
        "sexo",
        "fecha_nacimiento_fmt",
        "edad_admin",
        "etapa_admin",
        "raza",
        "color",
        "finca",
        "lote",
        "productor",
        "estado",
        "peso_nacimiento",
        "peso_destete",
        "created_at",
        "updated_at",
    )

    search_fields = (
        "id_interno",
        "id_siniga",
        "nombre_bov",
        "raza__raza",
        "color__color",
        "finca__nombre",
        "lote__nombre",
    )

    list_filter = (
        "sexo",
        "estado",
        "raza",
        "color",
        "finca",
        "lote",
        SinigaVacioFilter,
        FechaNacimientoFilter,
        PesoNacimientoFilter,
        PesoDesteteFilter,
    )

    ordering = ("-id",)

    autocomplete_fields = (
        "productor",
        "upp",
        "color",
        "raza",
        "registro",
        "finca",
        "lote",
        "padre",
        "madre",
    )

    readonly_fields = (
        "fecha_nacimiento_fmt",
        "edad_admin",
        "etapa_admin",
        "created_at",
        "updated_at",
    )

    actions = [
        set_estado("ACTIVO"),
        set_estado("VENDIDO"),
        set_estado("MUERTO"),
        set_estado("BAJA"),
        export_animales_xlsx,
        imprimir_reporte_animales_pdf,
    ]

    fieldsets = (
        ("Identificación", {"fields": ("id_interno", "id_siniga", "nombre_bov", "sexo")}),
        ("Origen / Ubicación", {"fields": ("finca", "lote", "productor", "upp")}),
        ("Características", {"fields": ("raza", "color", "estado", "notas")}),
        ("Pesos / Fechas", {"fields": ("fecha_nacimiento", "fecha_nacimiento_fmt", "peso_nacimiento", "peso_destete")}),
        ("Edad / Etapa (calculado)", {"fields": ("edad_admin", "etapa_admin")}),
        ("Genealogía", {"fields": ("padre", "madre", "registro")}),
        ("Sistema", {"fields": ("created_at", "updated_at")}),
    )

    @admin.display(description="Fecha nacimiento", ordering="fecha_nacimiento")
    def fecha_nacimiento_fmt(self, obj):
        return obj.fecha_nacimiento.strftime("%d/%m/%Y") if obj.fecha_nacimiento else "—"

    @admin.display(description="ID SINIGA", ordering="id_siniga")
    def id_siniga_admin(self, obj):
        return obj.id_siniga or "Sin Siniga"

    @admin.display(description="Edad (Años, Meses, Días)", ordering="fecha_nacimiento")
    def edad_admin(self, obj):
        return str(getattr(obj, "edad", "") or "")

    @admin.display(description="Etapa de desarrollo")
    def etapa_admin(self, obj):
        return str(getattr(obj, "etapa_desarrollo", "") or "")

    def _peso_warnings(self, obj):
        warnings = []
        if obj.sexo and obj.peso_nacimiento is not None:
            if obj.sexo == "M":
                if obj.peso_nacimiento < Decimal("30"):
                    warnings.append("Advertencia: peso nacimiento bajo (<30 kg) para macho.")
                elif obj.peso_nacimiento > Decimal("45"):
                    warnings.append("Advertencia: peso nacimiento alto (>45 kg) para macho.")
            elif obj.sexo == "H":
                if obj.peso_nacimiento < Decimal("28"):
                    warnings.append("Advertencia: peso nacimiento bajo (<28 kg) para hembra.")
                elif obj.peso_nacimiento > Decimal("42"):
                    warnings.append("Advertencia: peso nacimiento alto (>42 kg) para hembra.")

        if obj.sexo and obj.peso_destete is not None:
            if obj.sexo == "M":
                if obj.peso_destete < Decimal("160"):
                    warnings.append("Advertencia: peso destete bajo (<160 kg) para macho.")
                elif obj.peso_destete > Decimal("260"):
                    warnings.append("Advertencia: peso destete alto (>260 kg) para macho.")
            elif obj.sexo == "H":
                if obj.peso_destete < Decimal("150"):
                    warnings.append("Advertencia: peso destete bajo (<150 kg) para hembra.")
                elif obj.peso_destete > Decimal("240"):
                    warnings.append("Advertencia: peso destete alto (>240 kg) para hembra.")

        return warnings

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        for msg in self._peso_warnings(obj):
            messages.warning(request, msg)
